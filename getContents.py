import math
import time
import requests
import asyncio
import json
import datetime
import gspread

from datetime import date
from openpyxl.styles import Alignment
from googleapiclient.discovery import build  # Added
from google.oauth2 import service_account

from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from gspread_formatting import batch_updater

class ContentSCraping:
    # Init
    def __init__(self, month, day, event_url):
        self.month = month
        self.day = day
        self.event_url = event_url

    # Get Data from purpose site
    async def scanData(self):
        # Send get request
        async def send_request(p_url):
            url = p_url
            headers = {
                'Content-Type': 'application/json',
                "Accept": "application/json",
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36',
            }
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                json_response = response.text
                return json_response
            else:
                return ''

        # Append to array
        async def append_to_arr(store, data):
            length = len(store)
            if(length == 0):
                length = 1
            for i in range(len(data)):
                store.append([length + i, data[i]['userInfo']['displayName'], data[i]['score']])
            return store

        # Get Ranking List
        async def getRankingList(containerID):
            result = []
            url = f'https://api-dsa.17app.co/api/v1/leaderboards/eventory?containerID={containerID}&cursor=&count=100'
            nextCursor = 'init'
            while nextCursor != '':
                ranking_list = await send_request(url)
                ranking_list = json.loads(ranking_list)
                if 'data' in ranking_list:
                    await append_to_arr(result, ranking_list['data'])
                nextCursor = ranking_list['nextCursor']
                url = f'https://api-dsa.17app.co/api/v1/leaderboards/eventory?containerID={containerID}&cursor={nextCursor}&count=100'

            return result

        # Create New Google Sheet
        async def createGoogleSheet(filename):
            SCOPES = ['https://www.googleapis.com/auth/drive']  # Modified
            credentials = service_account.Credentials.from_service_account_file('service-account.json', scopes=SCOPES)

            drive = build('drive', 'v3', credentials=credentials)
            file_metadata = {
                'name': filename,
                'parents': ['11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM'],
                'mimeType': 'application/vnd.google-apps.spreadsheet'
            }
            res = drive.files().create(body=file_metadata).execute()
            permission_body = {
                'role': 'writer',  # Set the desired role ('reader', 'writer', 'commenter', 'owner')
                'type': 'anyone',  # Share with anyone
            }
            drive.permissions().create(fileId=res['id'], body=permission_body).execute()

            return res['id']

        # Get Sheet ID by file name in special folder
        async def get_sheet_by_name(file_name, folder_name):
            SCOPES = ['https://www.googleapis.com/auth/drive']  # Modified
            credentials = service_account.Credentials.from_service_account_file('service-account.json', scopes=SCOPES)
            drive_service = build('drive', 'v3', credentials=credentials)
            sheet_id = None

            results = drive_service.files().list(q="name='" + file_name + "' and and mimeType='application/vnd.google-apps.spreadsheet' ",
                                    pageSize=10, fields="nextPageToken, files(id, name)").execute()
            items = results.get('files', [])
            if not items:
                return ''
            else:
                sheet_id = items[0]['id']

            return sheet_id

        # Get Calculate result
        def calculate_date(year, month, day):
            date1 = datetime.datetime.strptime(f'{year}-{month}-{day}', '%Y-%m-%d')
            if(int(self.month) > month):
                start_year = year - 1
            else:
                start_year = year
            date2 = datetime.datetime.strptime(f'{start_year}-{self.month}-{self.day}', '%Y-%m-%d')
            delta = date1 - date2
            return delta.days
        
        # Create new sheet into spreadsheet
        async def create_sheet_into_spreadsheet(sheetID, data):
            SCOPES = ['https://www.googleapis.com/auth/drive']
            SERVICE_ACCOUNT_FILE = 'service-account.json'

            creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_key(sheetID)

            data = data['Data']
            for i in range(len(data)):
                worksheet = spreadsheet.add_worksheet(title=data[i]['EventID'], rows='500', cols='100')

        # Insert image into worksheet
        async def insert_image_in_googlesheet(sheetID, image):
            SCOPES = ['https://www.googleapis.com/auth/drive']
            SERVICE_ACCOUNT_FILE = 'service-account.json'

            creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_key(sheetID)

            sheet = spreadsheet.sheet1
            sheet.update_title('タイトル')

            sheet_range = "A1"
            batch = batch_updater(sheet.spreadsheet)
            batch.set_row_height(sheet, '1:1', 440)
            batch.set_column_width(sheet, 'A:A', 500)
            batch.execute()
            
            insert_image = f"=IMAGE(\"{image}\", 1)"
            sheet.update(sheet_range, [[insert_image]], value_input_option="USER_ENTERED")

        # Insert html content into worksheet
        async def insert_content_in_googlesheet(sheetID, element, parent_title, title):
            SCOPES = ['https://www.googleapis.com/auth/drive']
            SERVICE_ACCOUNT_FILE = 'service-account.json'

            creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_key(sheetID)
            
            worksheet = spreadsheet.add_worksheet(title=f"{parent_title} - {title}", rows='500', cols='100')

            search_panel = element.find_elements('css selector', '.bpEaZC')
            if(len(search_panel) > 0):
                return 

            i = 1
            contents = element.find_elements('css selector', '.bjzlAe')
            if(len(contents) == 0):
                contents = element.find_elements('css selector', '.btCdvi')
 
            for content in contents:
                child_elements = content.find_elements(By.TAG_NAME, '*')
                
                if(len(child_elements) > 0):
                    for child in child_elements:
                        class_name = child.get_attribute('class')[10:]
                        tag_name = child.tag_name

                        if(tag_name == 'center'):
                            worksheet.update(f"A{i}", [[child.text]], value_input_option="USER_ENTERED")
                            i += 1
                        elif(class_name == 'hCXNzI'):
                            worksheet.update(f"A{i}", [[child.text]], value_input_option="USER_ENTERED")
                            i += 1
                        elif class_name == 'jPbYFU' or class_name == 'fezHWk':
                            worksheet.update(f"A{i}", [[child.text]], value_input_option="USER_ENTERED")
                            i += 1
                        elif class_name == 'fpiBVx':
                            insert_image = f"=IMAGE(\"{child.get_attribute('src')}\", 1)"
                            batch = batch_updater(worksheet.spreadsheet)
                            batch.set_row_height(worksheet, f'1:{i}', 200)
                            batch.set_column_width(worksheet, f'A:A', 500)
                            batch.execute()

                            worksheet.update(f"A{i}", [[insert_image]], value_input_option="USER_ENTERED")
                            i += 1
                        elif class_name == 'dMxtIb':
                            worksheet.update(f"A{i}", [[child.text]], value_input_option="USER_ENTERED")
                            i += 1
                        elif class_name == 'bsffay':
                            worksheet.update(f"A{i}", [[child.text]], value_input_option="USER_ENTERED")
                            image_elements = child.find_elements(By.TAG_NAME, 'img')
                            i += 1

                            if(len(image_elements) > 0):
                                for image in image_elements:
                                    batch = batch_updater(worksheet.spreadsheet)
                                    batch.set_row_height(worksheet, f'{i}:1', 200)
                                    batch.set_column_width(worksheet, f'A{i}:A1', 500)
                                    batch.execute()

                                    insert_image = f"=IMAGE(\"{image.get_attribute('src')}\", 1)"
                                    worksheet.update(f"A{i}", [[insert_image]], value_input_option="USER_ENTERED")
                                    i += 1

                        elif class_name == 'bXAnVj':
                            th_element = child.find_elements(By.CLASS_NAME, 'jwdikc')
                            col_cnt = len(th_element)
                            data = []
                            td_element = child.find_elements(By.CLASS_NAME, 'cdkoph')
                            for td in td_element:
                                data.append(td.text)

                            k = 0
                            res_str = ''
                            for k in range(len(data)):
                                j = math.floor(k % col_cnt)
                                if(j == 0):
                                    res_str += data[k]
                                else:
                                    res_str += f" | {data[k]}"
                                
                                if(j == col_cnt - 1):
                                    worksheet.update(f"A{i}", [[res_str]], value_input_option="USER_ENTERED")
                                    res_str = ''

                                i += 1
                                print(f"----------------{i}-------------------")

                        print(f"----------------{i}-------------------")

        # Get attr of element
        async def handleGetAttr(elements, type):
            for element in elements:
                url = element.get_attribute(type)
                return url

        # Insert jpg, html content into spreadsheet
        async def insert_image(sheetID, event_id):
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")

            browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)


            # Maximize the browser window
            # browser.maximize_window()

            # Bring the browser window to the front
            # browser.execute_script("window.focus();")
            
            browser.get(f'https://event.17.live/{event_id}')
            time.sleep(10)
            main_image = ''
            main_image_elements = browser.find_elements('css selector', '.sc-crHlIS')
            main_video_elements = browser.find_elements('css selector', '.diUfYd')

            if(len(main_image_elements) > 0):
                main_image = await handleGetAttr(main_image_elements, 'src')
            else:
                main_image = await handleGetAttr(main_video_elements, 'src')
            
            print(main_image)
            await insert_image_in_googlesheet(sheetID, main_image)

            tab_elements = browser.find_elements('css selector', '.kGvAFP')
            for i in range(len(tab_elements)):
                if(i == 0):
                    continue
                tab_title = tab_elements[i].text
                print(tab_title)
                print('================')
                tab_elements[i].click()

                sub_tab_group = browser.find_elements('css selector', '.gOMukq')
                if(len(sub_tab_group) > 0):
                    sub_tab_elements = sub_tab_group[0].find_elements('css selector', '.ffjCOc')
                    if(len(sub_tab_elements) > 0):
                        for j in range(len(sub_tab_elements)):
                            sub_tab_title = sub_tab_elements[j].text
                            sub_tab_elements[j].click()
                            # 
                            last_sub_tab_group = browser.find_elements('css selector', '.gOMukq')
                            if(len(last_sub_tab_group) > 1):
                                last_sub_tab_elements = last_sub_tab_group[1].find_elements('css selector', '.ffjCOc')
                                if(len(last_sub_tab_elements) > 0):
                                    for l in range(len(last_sub_tab_elements)):
                                        sub_tab_title = last_sub_tab_elements[l].text
                                        last_sub_tab_elements[l].click()
                                        time.sleep(30)
                                        await insert_content_in_googlesheet(sheetID, browser, tab_title, sub_tab_title)
                            # 
                            time.sleep(30)
                            await insert_content_in_googlesheet(sheetID, browser, tab_title, sub_tab_title)

        # Write content into google sheets
        def write_into_googlesheet(sheetID, data):
            SCOPES = ['https://www.googleapis.com/auth/drive']
            SERVICE_ACCOUNT_FILE = 'service-account.json'

            creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_key(sheetID)

            data = data['Data']
            for i in range(len(data)):
                worksheet = spreadsheet.worksheet(data[i]['EventID'])
                start_row = 1  # Replace with the starting row index
                start_col = 2  # Replace with the starting column index
                end_row = 1    # Replace with the ending row index
                end_col = 3    # Replace with the ending column index

                # Merge the cells
                worksheet.merge_cells(start_row, start_col, end_row, end_col)

                # Write text in the merged cell
                merged_cell = worksheet.cell(start_row, start_col)  # Use the top-left cell of the merged range

                current_month = datetime.datetime.now().month
                current_day = datetime.datetime.now().day

                merged_cell.value = f"{current_month}月{current_day}日"
                merged_cell.alignment = Alignment(horizontal='center')

                # Update the sheet with the modified cell
                worksheet.update_cells([merged_cell])

                row_index = 2  # Assuming you want to insert the data in the 2nd row

                worksheet.insert_rows(data[i]['List'], row=row_index)
                time.sleep(10)

        # Get event all url
        # url = 'https://wap-api.17app.co/api/v1/event?region=JP&status=1'

        # event_urls = None
        # try:
        #     event_urls = await send_request(url)
        # except Exception as e:
        #     print(f"An error occurred while fetching JSON for {url}: {e}")

        # event_urls = json.loads(event_urls)
        # event_urls = event_urls['events']['inProgress']

        # event_url_arr = []
        # if (len(event_urls) > 1):
        #     for i in range(len(event_urls)):
        #         event_url_arr.append(event_urls[i]['descriptionURL'])

        event_url_arr = [self.event_url]

        # Get containerID from event refrence api
        event_json_data = []
        for i in range(len(event_url_arr)):
            event_id = event_url_arr[i].split('/')[-1]
            print(event_id)
            json_url = f'https://webcdn.17app.co/campaign/projects/{event_id}/references.json'
            try:
                event_json_response = await send_request(json_url)
                if(event_json_response != ''):
                    data = json.loads(event_json_response)
                    data = data['fetcher']
                    # Get Current Date
                    current_date = date.today()
                    formatted_date = current_date.strftime("%Y-%m-%d")

                    # Get event data.
                    event_data = []
                    for i in range(len(data)):
                        event_data.append({
                            "EventID": data[i]['id'][12:],
                            "ContainerID": data[i]['value']['args'][0],
                            "List": []
                        })

                    res = {
                        "ID": event_id,
                        "Date": formatted_date,
                        "Data": event_data,
                        "Count": 0
                    }
                    event_json_data.append(res)
                else:
                    return 'Failure'
            except Exception as e:
                print(f"An error occurred while fetching JSON for {json_url}: {e}") 

        for i in range(len(event_json_data)):
            data = event_json_data[i]['Data']
            for j in range(len(data)):
                event_json_data[i]['Data'][j]['List'] = await getRankingList(data[j]['ContainerID'])

        for i in range(len(event_json_data)):
            current_year = datetime.datetime.now().year
            current_month = datetime.datetime.now().month
            current_day = datetime.datetime.now().day

            filename = f"Ranking_{event_json_data[i]['ID']}_{current_year}_{self.month}_{self.day}"

            if((int(self.month) == current_month and int(self.day) == current_day) or (int(self.month) < current_month and int(self.day) < current_day)):
                # create new sheet
                sheetID = await createGoogleSheet(filename)
                print(sheetID)
                event_json_data[i]['Count'] == 0
                await insert_image(sheetID, event_json_data[i]['ID'])
                await create_sheet_into_spreadsheet(sheetID, event_json_data[i])
                time.sleep(20)
                write_into_googlesheet(sheetID, event_json_data[i])
            else:
                folder_name = '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM'
                sheetID = await get_sheet_by_name(filename, folder_name)
                print(sheetID)
                event_json_data[i]['Count'] = calculate_date(current_year, current_month, current_day)
                time.sleep(20)
                write_into_googlesheet(sheetID, event_json_data[i])

        return event_json_data

    async def main(self):
        result = await self.scanData()
        return result

    # Run the main coroutine
    def run(self):
        result = asyncio.run(self.main())
        return result
